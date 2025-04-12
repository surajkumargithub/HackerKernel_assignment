from django.views.generic import CreateView, ListView, View
from django.shortcuts import render
from django.urls import reverse_lazy
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'form_template.html'
    success_url = reverse_lazy('author-list')

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'form_template.html'
    success_url = reverse_lazy('book-list')

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'form_template.html'
    success_url = reverse_lazy('borrow-list')

class AuthorListView(ListView):
    model = Author
    template_name = 'list_template.html'
    paginate_by = 5

class BookListView(ListView):
    model = Book
    template_name = 'list_template.html'
    paginate_by = 5

class BorrowListView(ListView):
    model = BorrowRecord
    template_name = 'list_template.html'
    paginate_by = 5

class ExportExcelView(View):
    def get(self, request):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        wb = openpyxl.Workbook()
        
        # Authors
        ws1 = wb.active
        ws1.title = "Authors"
        ws1.append(['ID', 'Name', 'Email', 'Bio'])
        for author in Author.objects.all():
            ws1.append([author.id, author.name, author.email, author.bio])

        # Books
        ws2 = wb.create_sheet(title="Books")
        ws2.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
        for book in Book.objects.all():
            ws2.append([book.id, book.title, book.genre, book.published_date, book.author.name])

        # BorrowRecords
        ws3 = wb.create_sheet(title="BorrowRecords")
        ws3.append(['ID', 'User Name', 'Book', 'Borrow Date', 'Return Date'])
        for record in BorrowRecord.objects.all():
            ws3.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

        wb.save(response)
        return response
